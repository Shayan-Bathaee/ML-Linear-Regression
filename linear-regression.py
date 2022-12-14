# Overview: This program was written to perform linear regresson on a data set to find the line of best fit. To do so, it uses 
# gradient descent.
# Creation Date: 1/13/2022
# Author: Shayan Bathaee
# Concepts were inspired by the following article: https://towardsdatascience.com/linear-regression-using-gradient-descent-97a6c8700931
# US GDP data is sourced from https://www.kaggle.com/datasets/alejopaullier/-gdp-by-country-1999-2022


import matplotlib.pyplot as plt
import matplotlib.animation as animation
import numpy as np
import sys
import openpyxl

# DEFINE FUNCTION TO GET X AND Y DATA FROM SPREADSHEET
def getDataFromSpreadsheet(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    global xname
    global yname
    xname = ws["A1"].value
    yname = ws["B1"].value
    xvalues = []
    yvalues = []
    for cell in ws["A"]:
        xvalues.append(cell.value)
    xvalues.pop(0)                          # pop the name
    for cell in ws["B"]:
        yvalues.append(cell.value)
    yvalues.pop(0)

    return {xname: xvalues, yname: yvalues} # dataDictionary input to the regression initializer function


# DEFINE CLASS TO PERFORM REGRESSION
class Regression:
    # initialize the regression with the specified iterations, learning rate, and data
    def __init__(self, limit, iterations, learningRate, dataDictionary):
        self.limit = limit
        self.iterations = iterations
        self.learningRate = learningRate
        self.loss = 0
        self.X = np.array(dataDictionary[xname])
        self.Y = np.array(dataDictionary[yname])
        self.Xscaled = self.X - np.min(self.X)          # used for calculating predicted Y
        self.m = 0
        self.b = np.min(self.Y)                         # initial prediction of the line is y = (min y value)
        self.numDatapoints = len(self.X)                # thenumber of X Y pairs we hae
        self.predictedY = self.m*self.X + self.b        # this stores the data for the predicted line. It just starts as y = np.min(self.Y)

    # calculate mean squared error
    def calculateLoss(self):
        self.loss = 0                                               # initialize the loss to 0
        for i in range(self.numDatapoints):                         # for each data point
            self.loss += (self.Y[i] - (self.predictedY[i])) ** 2    # calculate the squared error (actual Y - predicted Y) ^ 2

        self.loss = self.loss / self.numDatapoints                  # divide the summation by number of datapoints to get average loss
        return self.loss

    def updatePrediction(self):
        self.predictedY = self.m*self.Xscaled + self.b              # the prediction has to be calculated using the scaled X values

    # use gradient descent to adjust m and b and minimize loss
    def applyGradientDescent(self):
        self.dm = 0                                                 # initialize derivatives to 0
        self.db = 0
        for i in range(self.numDatapoints):                         
            self.dm += (self.X[i]*(self.Y[i] - self.predictedY[i])) # summation part of derivative with respect to m
            self.db += (self.Y[i] - self.predictedY[i])             # summation part of derivative with respect to b
        self.dm = (-2/self.numDatapoints)*self.dm                   # scaling part of derivative with respect to m
        self.db = (-2/self.numDatapoints)*self.db                   # scaling part of derivative with respect to b
        self.m = self.m - self.learningRate*self.dm                 # adjust m and b using the derivative values and learning rate
        self.b = self.b - self.learningRate*self.db
        self.calculateLoss()
        self.updatePrediction()
    

# TAKE CONSOLE INPUT
if "-help" in sys.argv:
    print("\n\tUSAGE: python linear-regression.py")
    print("\tFLAGS:")
    print("\t\t-help        Display this message")
    print("\t\t-l           Limit the number of iterations to a specified value")
    print("\t\t-lr          Specify a learning rate")
    print("\t\t-i           Specify an input file")
    print("\tSee https://github.com/Shayan-Bathaee/ML-Linear-Regression for more information\n")
    exit()

if "-l" in sys.argv:
    iterations = int(sys.argv[sys.argv.index('-l') + 1])
    limit = True
else:
    iterations = 0
    limit = False

if "-lr" in sys.argv:
    learningRate = float(sys.argv[sys.argv.index('-lr') + 1])
else:
    learningRate = 0.000001

if "-i" in sys.argv:
    filename = str(sys.argv[sys.argv.index('-i') + 1])
    dataDictionary = getDataFromSpreadsheet(filename)
else:
    filename = "US_GDP_example.xlsx"
    dataDictionary = getDataFromSpreadsheet(filename)


# INITIALIZE THE REGRESSION, INITIALIZE THE PLOT
LR = Regression(limit, iterations, learningRate, dataDictionary)
fig, ax = plt.subplots()                # create a figure, make ax the only subplot
ax.grid()
dataText = ax.text(0.02, 0.8, 'm = 0\nb = ' + str(round(np.min(LR.Y), 2)) + '\nIterations = 0', transform=ax.transAxes, bbox=dict(facecolor='white', edgecolor='black'))
line, = ax.plot(LR.X, LR.predictedY)    # starting line is y = minimum y value
count = 0
consoleOutput = ""


# DEFINE FUNCTIONS USED FOR ANIMATION
# Initialize the animation
def init():
    dataText.set_text("m = 0\nb = " + str(round(np.min(LR.Y), 2)) + "\nIterations = 0")
    return line, dataText

# Function to execute for each frame
def animate(i):
    global count, consoleOutput
    displayString = "m = " + str(round(LR.m, 2)) + "\nb = " + str(round(LR.b - np.min(LR.Y), 2)) + "\nMSE = " + str(round(LR.loss,2)) + "\nIterations = " + str(i)
    if i == iterations and LR.limit == True:
        consoleOutput = displayString
        ani.event_source.stop()                         # stop if we reach our iteration count and limit is true
    LR.applyGradientDescent()                           # use gradient descent to update m and b
    line.set_ydata(LR.predictedY)                       # reset the line with our new m and b
    dataText.set_text(displayString)
    count += 1
    return line, dataText


# BEGIN ANIMATION AND PLOTTING
ani = animation.FuncAnimation(fig, animate, init_func=init, interval=1, blit=True)
plt.plot(LR.X, LR.Y, 'ro', mec='black')                                         # plot the data points
plt.ylim([np.min(LR.Y) - 1, np.max(LR.Y) + 0.5*(np.max(LR.Y) - np.min(LR.Y))])  # set the limits of the graph depending on our range (extra space added above for labels)
plt.xlim([np.min(LR.X) - 1, np.max(LR.X) + 1])
if xname:
    plt.xlabel(xname)
if yname:
    plt.ylabel(yname)
plt.show()

# print the final slope, y intercept, mean squared error, and iterations to console
if consoleOutput == "":
    consoleOutput = "m = " + str(round(LR.m, 2)) + "\nb = " + str(round(LR.b - np.min(LR.Y), 2)) + "\nMSE = " + str(round(LR.loss, 2)) + "\nIterations = " + str(count)
print(consoleOutput)
